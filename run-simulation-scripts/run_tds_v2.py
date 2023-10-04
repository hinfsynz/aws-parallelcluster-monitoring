import andes
import pandas as pd
from openpyxl import load_workbook
import sys
import os


if len(sys.argv) != 4:
    print('Error: Invalid arguments. Usage:\n\trun_tds.py case_index fault_bus_number fault_clearing_time\n')
    print('1: Kundur system\n2: IEEE14 bus system\n3: WECC system')
    sys.exit(1)

#print(sys.argv[1])

supported_systems = ['placeholder','kundur', 'ieee14', 'wecc']
selection = int(sys.argv[1])
if selection < 1 or selection > 3:
    print('Currently only support kundur, ieee14 and wecc system!') 
    sys.exit(1)

system_to_study = supported_systems[selection]
print('selected system: {0}'.format(system_to_study))
file_path_dict = {"kundur":"kundur/kundur_full.xlsx",
                  "ieee14":"ieee14/ieee14_fault.json",
                  "wecc": "wecc/wecc_full.xlsx"}

if system_to_study == 'ieee14':
    fault_bus_no = int(sys.argv[2])
    fault_tc = float(sys.argv[3])

    ss = andes.load(andes.get_case(file_path_dict['ieee14']),
                default_config=True, 
                setup=False)
    # change the fault bus
    ss.Fault.alter('bus',1, fault_bus_no)
    # change contingency definition
    ss.Fault.alter('tc', 1, fault_tc)
    # change output file names
    ss.files.txt = '{0}_fault_bus{1}_tc{2}_out.txt'.format(system_to_study, fault_bus_no, fault_tc)
    ss.files.lst = '{0}_fault_bus{1}_tc{2}_out.lst'.format(system_to_study, fault_bus_no, fault_tc)
    ss.files.nps = '{0}_fault_bus{1}_tc{2}_out.nps'.format(system_to_study, fault_bus_no, fault_tc)
else:
    open_line_no = int(sys.argv[2])
    open_line_time = float(sys.argv[3])
    print('Appending a toggler sheet to the workbook')
    df = pd.DataFrame({'uid':[0],'idx':[1],'u':[1],'name':['Toggler_1'],'model':['Line'],'dev':['Line_{0}'.format(open_line_no)],'t':[open_line_time]})
    file_abs_path = os.path.join(andes.utils.paths.cases_root(), file_path_dict[system_to_study])
    print(file_abs_path)
    book = load_workbook(file_abs_path)
    if 'Toggler' in book.sheetnames:
        print('Toggler sheet already exists')
    else:
        #create a new copy of the spreadsheet with different toggle params
        new_file_path = os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), 'slurm_workdir',
                            '{0}_line_{1}_open_at_{2:5.3f}s_full.xlsx'.format(system_to_study, open_line_no, open_line_time))
        writer = pd.ExcelWriter(new_file_path, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name='Toggler', index=False)
        writer.close()
        print('Done appending. Load the data in Andes')
    ss = andes.load(andes.get_case(new_file_path),
                default_config=True, 
                setup=False)
    ss.files.txt = '{0}_line_{1}_open_at_{2:5.3f}s_out.txt'.format(system_to_study, open_line_no, open_line_time)
    ss.files.lst = '{0}_line_{1}_open_at_{2:5.3f}s_out.lst'.format(system_to_study, open_line_no, open_line_time)
    ss.files.nps = '{0}_line_{1}_open_at_{2:5.3f}s_out.nps'.format(system_to_study, open_line_no, open_line_time)
# complete case change
ss.setup()
# solve power flow
ss.PFlow.run()
# run time-domain simulations
ss.TDS.run()

